import os
import json
import re
import openai
import zipfile
from pymongo import MongoClient
from docx import Document
from openpyxl import load_workbook
from tempfile import TemporaryDirectory
from lxml import etree

openai.api_key = "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"

client = MongoClient("mongodb+srv://smocookie:smocookie@cluster0.btwrt.mongodb.net/")
db = client["personal_info_db"]
collection = db["detected_info"]

patterns = {
    "주민등록번호": r"\b\d{6}-\d{7}\b",
    "전화번호": r"\b010-\d{3,4}-\d{4}\b",
    "생년월일": r"\b\d{4}[-/]\d{2}[-/]\d{2}\b",
    "계좌번호": r"\b(?!010|19\d{2}|20\d{2})\d{4}-\d{2,4}-\d{2,5}\b",
    "여권번호": r"\b[A-Z]{1}\d{8}\b",
    "이메일": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    "카드번호": r"\b\d{4}-\d{4}-\d{4}-\d{4}\b"
    # "주소": r"\b[가-힣]{2,}(시|도)\s[가-힣]{2,}(구|군)\s[가-힣0-9]{1,}(동|로|길)\s*\d*[-\d]*\b"
}


def detect_pii_with_regex(content):
    results = {}
    for key, pattern in patterns.items():
        matches = re.findall(pattern, content)
        if matches:
            results[key] = list(set(matches))  # 중복 제거
    return results


def detect_sensitive_info_with_chatgpt(content, additional_info):
    prompt = f"""
    다음 텍스트에서 개인정보(이름 및 주소)와 추가 요청된 정보를 탐지해주세요:
    - 개인정보에는 이름, 주소, 이메일, 전화번호, 생년월일, 계좌번호, 카드번호 등 개인을 특정할 수 있는 정보가 포함됩니다.
    - 추가 요청 정보: {additional_info}
    반환 형식(JSON):
    {{
        "개인정보": {{
            "이름": [],
            "주소": [],
            "전화번호": []
        }},
        "추가 탐지 정보": {{
            "추가 요청 정보": []
        }}
    }}
    텍스트:
    {content}
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )

     # API 응답 출력 (디버깅용)
    print("📊 [API 탐지 결과] -------------------")
    print(response.choices[0].message.content)
    
    try:
        return json.loads(response['choices'][0]['message']['content'])
    except json.JSONDecodeError:
        return {"error": "Invalid JSON from ChatGPT"}


def save_to_mongodb(file_name, detected_info, additional_results):
    document = {
        "file_name": file_name,
        "detected_info": detected_info,
        "chatgpt_plus_results": additional_results
    }
    collection.insert_one(document)


def get_masking_data_from_mongodb():
    documents = collection.find({})
    masking_data = set()
    for doc in documents:
        for key, values in doc.get("detected_info", {}).items():
            masking_data.update(values)
    return masking_data

# ✨ 정규표현식을 이용한 마스킹 적용
def apply_masking(content, masking_data):
    for item in masking_data:
        content = content.replace(item, "****")
        # content = re.sub(re.escape(item), "****", content)
    return content

# 📂 Word 문서 텍스트 추출
def extract_text_from_word(file_path):
    document = Document(file_path)
    return "\n".join([paragraph.text for paragraph in document.paragraphs])

# 📂 Excel 문서 텍스트 추출
def extract_text_from_excel(file_path):
    workbook = load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
    return text

# 🔄 XML 기반 마스킹 적용
def process_xml_file(xml_path, masking_data):
    parser = etree.XMLParser(remove_blank_text=True)
    with open(xml_path, 'rb') as file:
        xml_tree = etree.parse(file, parser)

    for element in xml_tree.xpath(".//w:t", namespaces={"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
        if element.text:
            element.text = apply_masking(element.text, masking_data)

    with open(xml_path, 'wb') as file:
        file.write(etree.tostring(xml_tree, pretty_print=True, xml_declaration=True, encoding='utf-8'))

# 📄 Word 파일 마스킹
def mask_sensitive_data_with_images(file_path):
    masking_data = get_masking_data_from_mongodb()
    
    print(f"🔒 마스킹 대상 정보: {masking_data}")
    
    with TemporaryDirectory() as temp_dir:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        document_xml_path = os.path.join(temp_dir, "word", "document.xml")
        if os.path.exists(document_xml_path):
            process_xml_file(document_xml_path, masking_data)

        new_file_path = file_path.replace(".docx", "(masked).docx")
        with zipfile.ZipFile(new_file_path, 'w') as zip_out:
            for foldername, subfolders, filenames in os.walk(temp_dir):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_out.write(file_path, arcname)

    return new_file_path

# 🎯 메인 실행 함수
def main(file_path, file_type, additional_info):
    if file_type == "word":
        content = extract_text_from_word(file_path)
    elif file_type == "excel":
        content = extract_text_from_excel(file_path)
    else:
        print("지원하지 않는 파일 형식입니다.")
        return

    regex_results = detect_pii_with_regex(content)
    chatgpt_response = detect_sensitive_info_with_chatgpt(content, additional_info)

    if "error" in chatgpt_response:
        print("ChatGPT 탐지 중 오류 발생:", chatgpt_response["error"])
        return

    chatgpt_results = chatgpt_response.get("개인정보", {})
    additional_results = chatgpt_response.get("추가 탐지 정보", {})

    final_results = {**regex_results, **chatgpt_results}
    print(f"🔒 마스킹 대상 정보1: {final_results}")

    save_to_mongodb(file_path, final_results, additional_results)

    masked_file = mask_sensitive_data_with_images(file_path)

    print(f"마스킹된 파일이 저장되었습니다: {masked_file}")

# 🔄 실행
if __name__ == "__main__":
    import sys
    file_path = sys.argv[1]
    file_type = sys.argv[2]
    additional_info = sys.argv[3]
    main(file_path, file_type, additional_info)
